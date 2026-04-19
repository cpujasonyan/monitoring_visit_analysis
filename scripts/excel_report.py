# -*- coding: utf-8 -*-
"""Monitoring Visit Analysis - Excel Report Generation."""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import FINALIZED, ARCHIVED_VAL
from analyze import is_overdue


# Styles
_hfont = Font(bold=True, color='FFFFFF', size=11)
_hfill = PatternFill('solid', fgColor='0D7377')
_od_fill = PatternFill('solid', fgColor='FFC7CE')
_od_font = Font(color='9C0006')
_warn_fill = PatternFill('solid', fgColor='FFEB9C')
_warn_font = Font(color='9C6500')
_good_fill = PatternFill('solid', fgColor='C6EFCE')
_good_font = Font(color='006100')
_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
_center = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _style_header(ws, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _hfont; cell.fill = _hfill; cell.alignment = _center; cell.border = _border


def _sc(ws, r, c, mode=None):
    cell = ws.cell(row=r, column=c)
    cell.border = _border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if mode == 'od': cell.fill = _od_fill; cell.font = _od_font
    elif mode == 'warn': cell.fill = _warn_fill; cell.font = _warn_font
    elif mode == 'good': cell.fill = _good_fill; cell.font = _good_font


def _auto_w(ws, n, maxr):
    for c in range(1, n + 1):
        mx = 0
        for r in range(1, min(maxr + 1, 50)):
            v = ws.cell(row=r, column=c).value
            if v: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(mx * 1.3 + 2, 8), 32)


def generate_excel(rows_data, visited, cra_stats, month_stats, site_stats, summary, output_path, today):
    """Generate the multi-sheet Excel analysis report."""
    owb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws5 = owb.active; ws5.title = '\u6c47\u603b\u6982\u89c8'
    sm = summary
    data = [
        ['Monitoring Visit \u5206\u6790\u62a5\u544a', ''],
        ['\u6570\u636e\u622a\u6b62\u65e5\u671f', str(today)], [''],
        ['\u6307\u6807', '\u6570\u503c', '\u8bf4\u660e'],
        ['\u8bbf\u89c6\u8bb0\u5f55\u603b\u6570', sm['total_records'], ''],
        ['\u5df2\u5b8c\u6210\u8bbf\u89c6\u6570', sm['total_visited'], ''],
        ['\u5df2\u9012\u4ea4\u62a5\u544a\u6570', sm['total_submitted'], ''],
        ['\u5df2\u5b9a\u7a3f\u62a5\u544a\u6570', sm['total_finalized'], ''],
        ['\u5df2\u5f52\u6863\u6570', sm['total_archived'],
         f"\u5f52\u6863\u7387 {sm['total_archived']/sm['total_visited']*100:.1f}%" if sm['total_visited'] else ''],
        [''], ['\u8d85\u671f\u7edf\u8ba1'],
        ['\u9012\u4ea4\u8d85\u671f\u6b21\u6570', sm['total_sub_od'],
         f"\u5360\u6bd4 {sm['total_sub_od']/sm['total_visited']*100:.1f}%" if sm['total_visited'] else ''],
        ['\u5b9a\u7a3f\u8d85\u671f\u6b21\u6570', sm['total_fin_od'],
         f"\u5360\u6bd4 {sm['total_fin_od']/sm['total_visited']*100:.1f}%" if sm['total_visited'] else ''],
        ['\u8ddf\u8fdb\u51fd\u8d85\u671f\u6b21\u6570', sm['total_fu_od'], '\u4ec5\u7edf\u8ba1\u5df2\u5b9a\u7a3f\u62a5\u544a'],
        ['\u7591\u4f3c\u6f0f\u767b\u8bb0\u6b21\u6570', sm['total_missing'], '\u5df2\u5b9a\u7a3f+\u8ddf\u8fdb\u51fd\u622a\u6b62\u5df2\u8fc7+\u672a\u53d1\u9001'],
        [''], ['\u98ce\u9669\u8bc6\u522b'],
        ['\u9ad8\u98ce\u9669CRA', len(sm['high_risk_cras']),
         ', '.join(sm['high_risk_cras']) if sm['high_risk_cras'] else '\u65e0'],
        ['\u9ad8\u98ce\u9669\u7814\u7a76\u4e2d\u5fc3', len(sm['high_risk_sites']),
         ', '.join(sm['high_risk_sites']) if sm['high_risk_sites'] else '\u65e0'],
        [''], ['\u5206\u6790\u89c4\u5219'],
        ['\u5de5\u4f5c\u65e5\u8ba1\u7b97', 'chinese_calendar \u5904\u7406\u6cd5\u5b9a\u8282\u5047\u65e5\u53ca\u8c03\u4f11'],
        ['\u9012\u4ea4\u8d85\u671f', '\u5b9e\u9645\u8bbf\u89c6\u6700\u540e\u4e00\u5929\u8d77 > 5\u4e2a\u5de5\u4f5c\u65e5'],
        ['\u5b9a\u7a3f\u8d85\u671f', '\u5b9e\u9645\u8bbf\u89c6\u6700\u540e\u4e00\u5929\u8d77 > 10\u4e2a\u5de5\u4f5c\u65e5'],
        ['\u8ddf\u8fdb\u51fd\u8d85\u671f', 'min(\u8bbf\u89c6\u7ed3\u675f+10WD, \u5b9a\u7a3f\u65e5\u671f+1WD), \u4ec5\u5df2\u5b9a\u7a3f'],
        ['\u6f0f\u767b\u8bb0', '\u5df2\u5b9a\u7a3f + \u8ddf\u8fdb\u51fd\u622a\u6b62\u5df2\u8fc7 + \u8ddf\u8fdb\u51fd\u672a\u53d1\u9001'],
        ['\u98ce\u9669\u7b49\u7ea7', '\u4efb\u4e00\u7ef4\u5ea6\u22652\u2192\u9ad8\u98ce\u9669, =1\u2192\u5173\u6ce8, 0\u2192\u6b63\u5e38'],
    ]
    for i, rd in enumerate(data, 1):
        for j, v in enumerate(rd, 1):
            ws5.cell(row=i, column=j, value=v)
    ws5.cell(row=1, column=1).font = Font(bold=True, size=16, color='0D7377')
    ws5.column_dimensions['A'].width = 22
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 45

    # Sheet 2: Visit detail
    ws1 = owb.create_sheet('\u8bbf\u89c6\u660e\u7ec6(\u8d85\u671f\u6807\u6ce8)')
    dh = ['\u4e2d\u5fc3\u7f16\u53f7', '\u4e2d\u5fc3\u540d\u79f0', '\u8bbf\u89c6ID', '\u8bbf\u89c6\u7c7b\u578b', '\u76d1\u67e5\u5458',
          '\u5b9e\u9645\u8bbf\u89c6\u65e5\u671f', '\u62a5\u544a\u72b6\u6001',
          '\u62a5\u544a\u9012\u4ea4\u65e5\u671f', '\u9012\u4ea4\u622a\u6b62\u65e5', '\u9012\u4ea4WD', '\u9012\u4ea4\u8d85\u671f',
          '\u62a5\u544a\u5b9a\u7a3f\u65e5\u671f', '\u5b9a\u7a3f\u622a\u6b62\u65e5', '\u5b9a\u7a3fWD', '\u5b9a\u7a3f\u8d85\u671f',
          '\u8ddf\u8fdb\u51fd\u53d1\u9001\u65e5\u671f', '\u8ddf\u8fdb\u51fd\u622a\u6b62\u65e5', '\u8ddf\u8fdb\u51fd\u8d85\u671f',
          '\u6f0f\u767b\u8bb0\u6807\u8bc6', '\u98ce\u9669\u6807\u8bb0', '\u5f52\u6863\u72b6\u6001']
    for i, h in enumerate(dh, 1):
        ws1.cell(row=1, column=i, value=h)
    _style_header(ws1, len(dh))
    dr = 2
    for r in rows_data:
        if not r['visit_end']:
            continue
        dims = [1 if is_overdue(r.get('submit_overdue')) else 0,
                1 if is_overdue(r.get('final_overdue')) else 0,
                1 if is_overdue(r.get('followup_overdue')) else 0,
                1 if r.get('missing_reg') else 0]
        row_risk = '\u5f02\u5e38' if any(d > 0 for d in dims) else ''
        vals = [r['site_no'], r['site_name'], r['visit_id'], r['visit_type'], r['cra'],
                ', '.join(str(d) for d in r['actual_dates']), r['report_status'],
                ', '.join(str(d) for d in r['report_submit_date']) if r['report_submit_date'] else '',
                str(r['submit_deadline']) if r['submit_deadline'] else '', r['submit_wd'], r['submit_overdue'],
                ', '.join(str(d) for d in r['report_final_date']) if r['report_final_date'] else '',
                str(r['final_deadline']) if r['final_deadline'] else '', r['final_wd'], r['final_overdue'],
                ', '.join(str(d) for d in r['followup_date']) if r['followup_date'] else '',
                str(r['followup_deadline']) if r['followup_deadline'] else '',
                r.get('followup_overdue', ''), r.get('missing_reg', ''), row_risk,
                '\u5df2\u5f52\u6863' if str(r.get('archive_status', '')) == ARCHIVED_VAL else '\u672a\u5f52\u6863']
        for i, v in enumerate(vals, 1):
            ws1.cell(row=dr, column=i, value=v)
        for c in range(1, len(dh) + 1):
            _sc(ws1, dr, c)
        if is_overdue(r.get('submit_overdue')):
            for c in [8, 9, 10, 11]: _sc(ws1, dr, c, 'od')
        elif r.get('submit_overdue') == '\u5426': _sc(ws1, dr, 11, 'good')
        if is_overdue(r.get('final_overdue')):
            for c in [12, 13, 14, 15]: _sc(ws1, dr, c, 'od')
        elif r.get('final_overdue') == '\u5426': _sc(ws1, dr, 15, 'good')
        if is_overdue(r.get('followup_overdue')):
            for c in [16, 17, 18]: _sc(ws1, dr, c, 'od')
        elif r.get('followup_overdue') == '\u5426': _sc(ws1, dr, 18, 'good')
        if r.get('missing_reg'): _sc(ws1, dr, 19, 'warn')
        dr += 1
    _auto_w(ws1, len(dh), dr)
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(dh))}{dr - 1}"
    ws1.freeze_panes = 'A2'

    # Sheet 3: CRA stats
    ws2 = owb.create_sheet('CRA\u7edf\u8ba1')
    ch = ['CRA', '\u5df2\u8bbf\u89c6', '\u5df2\u9012\u4ea4', '\u9012\u4ea4\u8d85\u671f', '\u5df2\u5b9a\u7a3f', '\u5b9a\u7a3f\u8d85\u671f',
          '\u8ddf\u8fdb\u51fd\u8d85\u671f', '\u6f0f\u767b\u8bb0', '\u5e73\u5747\u9012\u4ea4WD', '\u5e73\u5747\u5b9a\u7a3fWD',
          '\u5df2\u5f52\u6863', '\u5f52\u6863\u7387', '\u98ce\u9669\u7b49\u7ea7']
    for i, h in enumerate(ch, 1):
        ws2.cell(row=1, column=i, value=h)
    _style_header(ws2, len(ch))
    cr = 2
    for name in sorted(cra_stats, key=lambda x: cra_stats[x]['total'], reverse=True):
        s = cra_stats[name]
        vals = [name, s['total'], len(s['sub_wd']), s['sub_od'], len(s['fin_wd']), s['fin_od'],
                s['fu_od'], s['missing'],
                f"{sum(s['sub_wd'])/len(s['sub_wd']):.1f}" if s['sub_wd'] else 'N/A',
                f"{sum(s['fin_wd'])/len(s['fin_wd']):.1f}" if s['fin_wd'] else 'N/A',
                s['archived'], f"{s['archived']/s['total']*100:.1f}%" if s['total'] else 'N/A', s['risk']]
        for i, v in enumerate(vals, 1):
            ws2.cell(row=cr, column=i, value=v)
        for c in range(1, len(ch) + 1): _sc(ws2, cr, c)
        if s['risk'] == '\u9ad8\u98ce\u9669': _sc(ws2, cr, len(ch), 'od')
        elif s['risk'] == '\u5173\u6ce8': _sc(ws2, cr, len(ch), 'warn')
        cr += 1
    _auto_w(ws2, len(ch), cr)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(ch))}{cr - 1}"
    ws2.freeze_panes = 'A2'

    # Sheet 4: Monthly stats
    ws3 = owb.create_sheet('\u6708\u5ea6\u7edf\u8ba1')
    mh = ['\u6708\u4efd', '\u8bbf\u89c6\u6b21\u6570', '\u5df2\u9012\u4ea4', '\u9012\u4ea4\u8d85\u671f', '\u9012\u4ea4\u8d85\u671f\u7387',
          '\u5df2\u5b9a\u7a3f', '\u5b9a\u7a3f\u8d85\u671f', '\u5b9a\u7a3f\u8d85\u671f\u7387',
          '\u8ddf\u8fdb\u51fd\u8d85\u671f', '\u8ddf\u8fdb\u51fd\u8d85\u671f\u7387', '\u6f0f\u767b\u8bb0',
          '\u5e73\u5747\u9012\u4ea4WD', '\u5e73\u5747\u5b9a\u7a3fWD', '\u5df2\u5f52\u6863', '\u5f52\u6863\u7387']
    for i, h in enumerate(mh, 1):
        ws3.cell(row=1, column=i, value=h)
    _style_header(ws3, len(mh))
    mr = 2
    for m in sorted(month_stats):
        s = month_stats[m]
        vals = [m, s['total'], s['submitted'], s['sub_od'],
                f"{s['sub_od']/s['submitted']*100:.1f}%" if s['submitted'] else 'N/A',
                s['finalized'], s['fin_od'],
                f"{s['fin_od']/s['finalized']*100:.1f}%" if s['finalized'] else 'N/A',
                s['fu_od'], f"{s['fu_od']/s['finalized']*100:.1f}%" if s['finalized'] else 'N/A',
                s['missing'],
                f"{sum(s['sub_wd'])/len(s['sub_wd']):.1f}" if s['sub_wd'] else 'N/A',
                f"{sum(s['fin_wd'])/len(s['fin_wd']):.1f}" if s['fin_wd'] else 'N/A',
                s['archived'], f"{s['archived']/s['total']*100:.1f}%" if s['total'] else 'N/A']
        for i, v in enumerate(vals, 1):
            ws3.cell(row=mr, column=i, value=v)
        for c in range(1, len(mh) + 1): _sc(ws3, mr, c)
        mr += 1
    _auto_w(ws3, len(mh), mr)
    ws3.freeze_panes = 'A2'

    # Sheet 5: Site stats
    ws4 = owb.create_sheet('Site\u7edf\u8ba1')
    sh = ['\u7f16\u53f7', '\u4e2d\u5fc3\u540d\u79f0', '\u5df2\u8bbf\u89c6', '\u5df2\u9012\u4ea4', '\u9012\u4ea4\u8d85\u671f',
          '\u5df2\u5b9a\u7a3f', '\u5b9a\u7a3f\u8d85\u671f', '\u8ddf\u8fdb\u51fd\u8d85\u671f', '\u6f0f\u767b\u8bb0',
          '\u5e73\u5747\u9012\u4ea4WD', '\u5e73\u5747\u5b9a\u7a3fWD', '\u5df2\u5f52\u6863', '\u5f52\u6863\u7387', '\u98ce\u9669\u7b49\u7ea7']
    for i, h in enumerate(sh, 1):
        ws4.cell(row=1, column=i, value=h)
    _style_header(ws4, len(sh))
    sr = 2
    for sid in sorted(site_stats, key=lambda x: site_stats[x]['total'], reverse=True):
        s = site_stats[sid]
        vals = [sid, s['name'], s['total'], len(s['sub_wd']), s['sub_od'], len(s['fin_wd']), s['fin_od'],
                s['fu_od'], s['missing'],
                f"{sum(s['sub_wd'])/len(s['sub_wd']):.1f}" if s['sub_wd'] else 'N/A',
                f"{sum(s['fin_wd'])/len(s['fin_wd']):.1f}" if s['fin_wd'] else 'N/A',
                s['archived'], f"{s['archived']/s['total']*100:.1f}%" if s['total'] else 'N/A', s['risk']]
        for i, v in enumerate(vals, 1):
            ws4.cell(row=sr, column=i, value=v)
        for c in range(1, len(sh) + 1): _sc(ws4, sr, c)
        if s['risk'] == '\u9ad8\u98ce\u9669': _sc(ws4, sr, len(sh), 'od')
        elif s['risk'] == '\u5173\u6ce8': _sc(ws4, sr, len(sh), 'warn')
        sr += 1
    _auto_w(ws4, len(sh), sr)
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(sh))}{sr - 1}"
    ws4.freeze_panes = 'A2'

    owb.save(output_path)
    return output_path
