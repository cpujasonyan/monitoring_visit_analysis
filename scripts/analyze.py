# -*- coding: utf-8 -*-
"""Monitoring Visit Analysis - Data Reading & Analysis Logic."""

import openpyxl
from collections import defaultdict
from config import (COLUMNS, SUBMIT_WD, FINAL_WD, FOLLOWUP_VISIT_WD, FOLLOWUP_FINAL_WD,
                    FINALIZED, FOLLOWUP_SENT, ARCHIVED_VAL, RISK_HIGH, RISK_ATTN)
from utils import parse_dates, last_date, first_date, working_days_between, add_working_days


def is_overdue(val):
    return '\u662f' in str(val or '')


def read_excel(filepath):
    """Read Monitoring Visit Tracking Log Excel and return list of record dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows_data = []
    C = COLUMNS
    for row in range(2, ws.max_row + 1):
        site_no = ws.cell(row=row, column=C['site_no']).value
        if site_no is None:
            continue
        rec = {
            'row': row,
            'site_no': str(site_no),
            'site_name': str(ws.cell(row=row, column=C['site_name']).value or ''),
            'visit_id': str(ws.cell(row=row, column=C['visit_id']).value or ''),
            'visit_type': str(ws.cell(row=row, column=C['visit_type']).value or ''),
            'cra': str(ws.cell(row=row, column=C['cra']).value or ''),
            'co_cra': str(ws.cell(row=row, column=C['co_cra']).value or ''),
            'actual_dates': parse_dates(ws.cell(row=row, column=C['actual_dates']).value),
            'report_status': str(ws.cell(row=row, column=C['report_status']).value or ''),
            'report_submit_date': parse_dates(ws.cell(row=row, column=C['report_submit_date']).value),
            'report_final_date': parse_dates(ws.cell(row=row, column=C['report_final_date']).value),
            'followup_status': str(ws.cell(row=row, column=C['followup_status']).value or ''),
            'followup_date': parse_dates(ws.cell(row=row, column=C['followup_date']).value),
            'archive_status': ws.cell(row=row, column=C['archive_status']).value,
        }
        rows_data.append(rec)
    wb.close()
    return rows_data


def compute_overdue(rows_data, today):
    """Compute overdue status for each record. Modifies records in-place.
    Returns list of visited records (those with visit_end date)."""
    for rec in rows_data:
        visit_end = last_date(rec['actual_dates'])
        submit_dt = first_date(rec['report_submit_date'])
        final_dt = first_date(rec['report_final_date'])
        followup_dt = first_date(rec['followup_date'])
        is_finalized = rec['report_status'] == FINALIZED
        rec['visit_end'] = visit_end

        # Submit deadline = visit_end + 5WD
        rec['submit_deadline'] = add_working_days(visit_end, SUBMIT_WD) if visit_end else None
        # Finalize deadline = visit_end + 10WD
        rec['final_deadline'] = add_working_days(visit_end, FINAL_WD) if visit_end else None

        # Submit overdue
        if visit_end and submit_dt:
            wd = working_days_between(visit_end, submit_dt)
            rec['submit_wd'] = wd
            rec['submit_overdue'] = '\u662f' if wd > SUBMIT_WD else '\u5426'
        elif visit_end and not submit_dt:
            wd = working_days_between(visit_end, today)
            rec['submit_wd'] = wd
            rec['submit_overdue'] = '\u662f(\u672a\u9012\u4ea4)' if wd > SUBMIT_WD else '\u672a\u9012\u4ea4'
        else:
            rec['submit_wd'] = None
            rec['submit_overdue'] = '\u672a\u8bbf\u89c6'

        # Finalize overdue
        if visit_end and final_dt:
            wd = working_days_between(visit_end, final_dt)
            rec['final_wd'] = wd
            rec['final_overdue'] = '\u662f' if wd > FINAL_WD else '\u5426'
        elif visit_end and not final_dt:
            wd = working_days_between(visit_end, today)
            rec['final_wd'] = wd
            rec['final_overdue'] = '\u662f(\u672a\u5b9a\u7a3f)' if wd > FINAL_WD else '\u672a\u5b9a\u7a3f'
        else:
            rec['final_wd'] = None
            rec['final_overdue'] = '\u672a\u8bbf\u89c6'

        # Follow-up overdue (only for finalized reports)
        rec['followup_deadline'] = None
        rec['followup_overdue'] = ''
        rec['followup_wd'] = None
        if visit_end and is_finalized and final_dt:
            dl_a = add_working_days(visit_end, FOLLOWUP_VISIT_WD)
            dl_b = add_working_days(final_dt, FOLLOWUP_FINAL_WD)
            rec['followup_deadline'] = min(dl_a, dl_b)
            if followup_dt:
                rec['followup_overdue'] = '\u662f' if followup_dt > rec['followup_deadline'] else '\u5426'
                rec['followup_wd'] = working_days_between(rec['followup_deadline'], followup_dt) if followup_dt > rec['followup_deadline'] else 0
            else:
                rec['followup_overdue'] = '\u662f(\u672a\u53d1\u9001)' if today > rec['followup_deadline'] else '\u672a\u53d1\u9001'
                rec['followup_wd'] = working_days_between(rec['followup_deadline'], today) if today > rec['followup_deadline'] else None

        # Missing registration (finalized + follow-up deadline passed + not sent)
        rec['missing_reg'] = ''
        if visit_end and is_finalized and final_dt and rec['followup_deadline']:
            followup_sent = (rec['followup_status'] == FOLLOWUP_SENT and followup_dt)
            if not followup_sent and today > rec['followup_deadline']:
                rec['missing_reg'] = '\u7591\u4f3c\u6f0f\u767b\u8bb0'

    return [r for r in rows_data if r['visit_end'] is not None]


def risk_level(sub_od, fin_od, fu_od, missing):
    dims = [sub_od, fin_od, fu_od, missing]
    if any(d >= RISK_HIGH for d in dims):
        return '\u9ad8\u98ce\u9669'
    if any(d >= RISK_ATTN for d in dims):
        return '\u5173\u6ce8'
    return '\u6b63\u5e38'


def compute_stats(visited):
    """Compute multi-dimensional statistics. Returns (cra_stats, month_stats, site_stats)."""
    cra_stats = defaultdict(lambda: {'total': 0, 'sub_od': 0, 'fin_od': 0, 'fu_od': 0,
        'missing': 0, 'sub_wd': [], 'fin_wd': [], 'archived': 0, 'not_sub': 0, 'od_details': []})
    month_stats = defaultdict(lambda: {'total': 0, 'sub_od': 0, 'fin_od': 0, 'fu_od': 0,
        'missing': 0, 'sub_wd': [], 'fin_wd': [], 'archived': 0, 'finalized': 0, 'submitted': 0})
    site_stats = defaultdict(lambda: {'name': '', 'total': 0, 'sub_od': 0, 'fin_od': 0,
        'fu_od': 0, 'missing': 0, 'sub_wd': [], 'fin_wd': [], 'archived': 0, 'od_details': []})

    for r in visited:
        # CRA
        s = cra_stats[r['cra']]
        s['total'] += 1
        if is_overdue(r.get('submit_overdue')): s['sub_od'] += 1
        if is_overdue(r.get('final_overdue')): s['fin_od'] += 1
        if is_overdue(r.get('followup_overdue')): s['fu_od'] += 1
        if r.get('missing_reg'): s['missing'] += 1
        if r.get('submit_wd') is not None and r['report_submit_date']: s['sub_wd'].append(r['submit_wd'])
        if r.get('final_wd') is not None and r['report_final_date']: s['fin_wd'].append(r['final_wd'])
        if str(r.get('archive_status', '')) == ARCHIVED_VAL: s['archived'] += 1
        if not r['report_submit_date'] and r['visit_end']: s['not_sub'] += 1
        details = []
        if is_overdue(r.get('submit_overdue')): details.append(f"\u9012\u4ea4\u8d85\u671f({r['submit_wd']}WD)")
        if is_overdue(r.get('final_overdue')): details.append(f"\u5b9a\u7a3f\u8d85\u671f({r['final_wd']}WD)")
        if is_overdue(r.get('followup_overdue')): details.append("\u8ddf\u8fdb\u51fd\u8d85\u671f")
        if r.get('missing_reg'): details.append("\u6f0f\u767b\u8bb0")
        if details:
            s['od_details'].append({'site': r['site_no'], 'visit': str(r['visit_end']), 'items': details})

        # Month
        m = r['visit_end'].strftime('%Y-%m')
        ms = month_stats[m]
        ms['total'] += 1
        if is_overdue(r.get('submit_overdue')): ms['sub_od'] += 1
        if is_overdue(r.get('final_overdue')): ms['fin_od'] += 1
        if is_overdue(r.get('followup_overdue')): ms['fu_od'] += 1
        if r.get('missing_reg'): ms['missing'] += 1
        if r.get('submit_wd') is not None and r['report_submit_date']: ms['sub_wd'].append(r['submit_wd'])
        if r.get('final_wd') is not None and r['report_final_date']: ms['fin_wd'].append(r['final_wd'])
        if str(r.get('archive_status', '')) == ARCHIVED_VAL: ms['archived'] += 1
        if r['report_final_date']: ms['finalized'] += 1
        if r['report_submit_date']: ms['submitted'] += 1

        # Site
        ss = site_stats[r['site_no']]
        ss['name'] = r['site_name']
        ss['total'] += 1
        if is_overdue(r.get('submit_overdue')): ss['sub_od'] += 1
        if is_overdue(r.get('final_overdue')): ss['fin_od'] += 1
        if is_overdue(r.get('followup_overdue')): ss['fu_od'] += 1
        if r.get('missing_reg'): ss['missing'] += 1
        if r.get('submit_wd') is not None and r['report_submit_date']: ss['sub_wd'].append(r['submit_wd'])
        if r.get('final_wd') is not None and r['report_final_date']: ss['fin_wd'].append(r['final_wd'])
        if str(r.get('archive_status', '')) == ARCHIVED_VAL: ss['archived'] += 1
        sd = []
        if is_overdue(r.get('submit_overdue')): sd.append("\u9012\u4ea4\u8d85\u671f")
        if is_overdue(r.get('final_overdue')): sd.append("\u5b9a\u7a3f\u8d85\u671f")
        if is_overdue(r.get('followup_overdue')): sd.append("\u8ddf\u8fdb\u51fd\u8d85\u671f")
        if r.get('missing_reg'): sd.append("\u6f0f\u767b\u8bb0")
        if sd:
            ss['od_details'].append({'cra': r['cra'], 'visit': str(r['visit_end']), 'items': sd})

    # Classify risk
    for name, s in cra_stats.items():
        s['risk'] = risk_level(s['sub_od'], s['fin_od'], s['fu_od'], s['missing'])
    for sid, s in site_stats.items():
        s['risk'] = risk_level(s['sub_od'], s['fin_od'], s['fu_od'], s['missing'])

    return dict(cra_stats), dict(month_stats), dict(site_stats)


def compute_summary(rows_data, visited, cra_stats, site_stats):
    """Compute summary statistics."""
    total_sub_od = sum(1 for r in visited if is_overdue(r.get('submit_overdue')))
    total_fin_od = sum(1 for r in visited if is_overdue(r.get('final_overdue')))
    total_fu_od = sum(1 for r in visited if is_overdue(r.get('followup_overdue')))
    total_missing = sum(1 for r in visited if r.get('missing_reg'))
    total_archived = sum(1 for r in visited if str(r.get('archive_status', '')) == ARCHIVED_VAL)
    high_risk_cras = [n for n, s in cra_stats.items() if s['risk'] == '\u9ad8\u98ce\u9669']
    high_risk_sites = [sid for sid, s in site_stats.items() if s['risk'] == '\u9ad8\u98ce\u9669']
    unique_cras = len(cra_stats)
    unique_sites = len(site_stats)
    return {
        'total_records': len(rows_data),
        'total_visited': len(visited),
        'total_submitted': sum(1 for r in visited if r['report_submit_date']),
        'total_finalized': sum(1 for r in visited if r['report_final_date']),
        'total_sub_od': total_sub_od,
        'total_fin_od': total_fin_od,
        'total_fu_od': total_fu_od,
        'total_missing': total_missing,
        'total_archived': total_archived,
        'high_risk_cras': high_risk_cras,
        'high_risk_sites': high_risk_sites,
        'unique_cras': unique_cras,
        'unique_sites': unique_sites,
    }
